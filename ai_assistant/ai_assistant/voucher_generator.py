from __future__ import annotations

import os
import re
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation

import frappe
from frappe.utils.file_manager import get_file_path, save_file

from .account_mapping import base_account, is_valid_account, map_transaction
from .excel_writer import build_financial_report_workbook

REQUIRED_HEADERS = ["入账日期", "转出金额", "转入金额", "余额", "对方单位", "对方账号", "摘要", "用途"]


def _to_amount(value):
    if value is None:
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0")


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _find_header_row(ws):
    max_rows = min(ws.max_row or 1, 30)
    for row_idx in range(1, max_rows + 1):
        values = [_clean(ws.cell(row=row_idx, column=col_idx).value) for col_idx in range(1, (ws.max_column or 1) + 1)]
        hits = sum(1 for header in REQUIRED_HEADERS if header in values)
        if hits >= 4:
            return row_idx, {header: values.index(header) + 1 for header in REQUIRED_HEADERS if header in values}
    raise ValueError("未找到银行交易明细表头。请确认文件包含：入账日期、转出金额、转入金额、余额、对方单位、摘要、用途等列。")


def parse_bank_transactions(source_path):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("当前运行环境缺少 openpyxl，无法读取 Excel。请在 bench 环境安装 openpyxl。") from exc

    wb = openpyxl.load_workbook(source_path, data_only=True, read_only=False)
    ws = wb.worksheets[0]
    header_row, columns = _find_header_row(ws)
    transactions = []

    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        row = {header: ws.cell(row=row_idx, column=col_idx).value for header, col_idx in columns.items()}
        out_amount = _to_amount(row.get("转出金额"))
        in_amount = _to_amount(row.get("转入金额"))
        if out_amount <= 0 and in_amount <= 0:
            continue

        transactions.append({
            "row_no": row_idx,
            "posting_date": row.get("入账日期"),
            "out_amount": out_amount,
            "in_amount": in_amount,
            "balance": _to_amount(row.get("余额")),
            "counterparty": _clean(row.get("对方单位")),
            "counterparty_account": _clean(row.get("对方账号")),
            "summary": _clean(row.get("摘要")),
            "purpose": _clean(row.get("用途")),
        })

    if not transactions:
        raise ValueError("Excel 中没有识别到有效的转入/转出交易记录。")
    return transactions


def _summary_for_transaction(tx):
    summary = tx.get("summary") or tx.get("purpose") or "银行交易"
    counterparty = tx.get("counterparty")
    if counterparty and counterparty not in summary and len(summary) < 18:
        return f"{summary}-{counterparty[:18]}"
    return summary


def _build_mapped_entries(transactions):
    mapped_entries = []
    review_notes = []
    candidates = []

    for idx, tx in enumerate(transactions, start=1):
        out_amount = tx["out_amount"]
        in_amount = tx["in_amount"]
        if out_amount > 0 and in_amount > 0:
            review_notes.append(f"第 {tx['row_no']} 行同时存在转入和转出金额，已按两笔方向分别处理。")

        directions = []
        if out_amount > 0:
            directions.append(("out", out_amount))
        if in_amount > 0:
            directions.append(("in", in_amount))

        for direction, amount in directions:
            debit_account, credit_account, reason = map_transaction(
                tx.get("summary"), tx.get("purpose"), tx.get("counterparty"), direction=direction
            )
            candidate_id = f"{idx}-{direction}-{tx['row_no']}"
            entry = {
                "candidate_id": candidate_id,
                "voucher_no": idx,
                "tx": tx,
                "direction": direction,
                "amount": amount,
                "debit_account": debit_account,
                "credit_account": credit_account,
                "mapping_reason": reason,
                "needs_review": "需人工复核" in reason,
                "ai_applied": False,
                "ai_confidence": None,
            }
            mapped_entries.append(entry)

            if entry["needs_review"]:
                candidates.append({
                    "id": candidate_id,
                    "source_row": tx["row_no"],
                    "direction": direction,
                    "amount": float(amount),
                    "posting_date": str(tx.get("posting_date") or ""),
                    "summary": tx.get("summary") or "",
                    "purpose": tx.get("purpose") or "",
                    "counterparty": tx.get("counterparty") or "",
                    "counterparty_account": tx.get("counterparty_account") or "",
                    "fallback_debit_account": debit_account,
                    "fallback_credit_account": credit_account,
                })

    return mapped_entries, candidates, review_notes


def _valid_ai_suggestion(direction, debit_account, credit_account):
    if not is_valid_account(debit_account) or not is_valid_account(credit_account):
        return False
    if direction == "out" and credit_account != "银行存款":
        return False
    if direction == "in" and debit_account != "银行存款":
        return False
    return True


def _apply_ai_suggestions(mapped_entries, ai_classifier=None):
    candidates = [
        {
            "id": entry["candidate_id"],
            "source_row": entry["tx"]["row_no"],
            "direction": entry["direction"],
            "amount": float(entry["amount"]),
            "posting_date": str(entry["tx"].get("posting_date") or ""),
            "summary": entry["tx"].get("summary") or "",
            "purpose": entry["tx"].get("purpose") or "",
            "counterparty": entry["tx"].get("counterparty") or "",
            "counterparty_account": entry["tx"].get("counterparty_account") or "",
            "fallback_debit_account": entry["debit_account"],
            "fallback_credit_account": entry["credit_account"],
        }
        for entry in mapped_entries
        if entry["needs_review"]
    ]
    if not candidates or not ai_classifier:
        return {"candidate_count": len(candidates), "applied_count": 0, "invalid_count": 0, "error": None}

    try:
        suggestions = ai_classifier(candidates) or {}
    except Exception as exc:
        return {"candidate_count": len(candidates), "applied_count": 0, "invalid_count": 0, "error": str(exc)}

    applied_count = 0
    invalid_count = 0
    for entry in mapped_entries:
        suggestion = suggestions.get(entry["candidate_id"])
        if not suggestion:
            continue

        debit_account = _clean(suggestion.get("debit_account"))
        credit_account = _clean(suggestion.get("credit_account"))
        if not _valid_ai_suggestion(entry["direction"], debit_account, credit_account):
            invalid_count += 1
            continue

        entry["debit_account"] = debit_account
        entry["credit_account"] = credit_account
        confidence = float(suggestion.get("confidence") or 0)
        entry["mapping_reason"] = f"AI辅助判断：{_clean(suggestion.get('reason')) or '未提供原因'}"
        entry["needs_review"] = confidence < 0.75
        entry["ai_applied"] = True
        entry["ai_confidence"] = confidence
        applied_count += 1

    return {"candidate_count": len(candidates), "applied_count": applied_count, "invalid_count": invalid_count, "error": None}


def build_voucher_rows(transactions, ai_classifier=None):
    voucher_rows = []
    mapped_entries, _candidates, review_notes = _build_mapped_entries(transactions)
    ai_stats = _apply_ai_suggestions(mapped_entries, ai_classifier=ai_classifier)

    for entry in mapped_entries:
        tx = entry["tx"]
        amount_float = float(entry["amount"])
        summary = _summary_for_transaction(tx)
        if entry["needs_review"]:
            review_notes.append(f"凭证 {entry['voucher_no']}（源文件第 {tx['row_no']} 行）需人工复核：{summary}；{entry['mapping_reason']}")

        voucher_rows.append({
            "voucher_no": entry["voucher_no"],
            "summary": summary,
            "debit_account": entry["debit_account"],
            "credit_account": None,
            "debit_amount": amount_float,
            "credit_amount": None,
            "source_row": tx["row_no"],
            "mapping_reason": entry["mapping_reason"],
            "needs_review": entry["needs_review"],
            "ai_applied": entry["ai_applied"],
            "ai_confidence": entry.get("ai_confidence"),
            "mapping_source": "AI" if entry["ai_applied"] else "规则",
            "review_status": "需复核" if entry["needs_review"] else "已匹配",
        })
        voucher_rows.append({
            "voucher_no": None,
            "summary": None,
            "debit_account": None,
            "credit_account": entry["credit_account"],
            "debit_amount": None,
            "credit_amount": amount_float,
            "source_row": tx["row_no"],
            "mapping_reason": entry["mapping_reason"],
            "needs_review": entry["needs_review"],
            "ai_applied": entry["ai_applied"],
            "ai_confidence": entry.get("ai_confidence"),
            "mapping_source": "AI" if entry["ai_applied"] else "规则",
            "review_status": "需复核" if entry["needs_review"] else "已匹配",
        })

    if ai_stats["error"]:
        review_notes.append(f"AI辅助判断失败，已使用规则兜底：{ai_stats['error']}")
    if ai_stats["invalid_count"]:
        review_notes.append(f"AI返回 {ai_stats['invalid_count']} 条不合法科目或方向，已丢弃并使用规则兜底。")

    return voucher_rows, review_notes, ai_stats


def _parse_posting_date(value):
    if isinstance(value, datetime):
        return value.date()
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def build_bank_balance_context(transactions):
    if not transactions:
        return {"opening_balance": 0.0, "ending_balance": 0.0, "has_bank_balance": False}

    # Bank statements are usually exported newest-first. For same-day rows,
    # larger source row numbers happened earlier in chronological order.
    chronological = sorted(
        transactions,
        key=lambda tx: (_parse_posting_date(tx.get("posting_date")) or datetime.min.date(), -(tx.get("row_no") or 0)),
    )
    first_tx = chronological[0]
    last_tx = chronological[-1]
    first_balance_after = first_tx.get("balance") or Decimal("0")
    opening_balance = first_balance_after - first_tx.get("in_amount", Decimal("0")) + first_tx.get("out_amount", Decimal("0"))
    ending_balance = last_tx.get("balance") or Decimal("0")

    return {
        "opening_balance": float(opening_balance),
        "ending_balance": float(ending_balance),
        "first_row_no": first_tx.get("row_no"),
        "last_row_no": last_tx.get("row_no"),
        "has_bank_balance": True,
    }


def build_trial_balance(voucher_rows, bank_context=None):
    totals = defaultdict(lambda: {"opening_debit": 0.0, "opening_credit": 0.0, "period_debit": 0.0, "period_credit": 0.0})

    bank_context = bank_context or {}
    opening_bank = float(bank_context.get("opening_balance") or 0)
    if bank_context.get("has_bank_balance") and opening_bank:
        if opening_bank > 0:
            totals["银行存款"]["opening_debit"] += opening_bank
            totals["利润分配"]["opening_credit"] += opening_bank
        else:
            totals["银行存款"]["opening_credit"] += abs(opening_bank)
            totals["利润分配"]["opening_debit"] += abs(opening_bank)

    for row in voucher_rows:
        debit_account = row.get("debit_account")
        credit_account = row.get("credit_account")
        if debit_account:
            totals[base_account(debit_account)]["period_debit"] += float(row.get("debit_amount") or 0)
        if credit_account:
            totals[base_account(credit_account)]["period_credit"] += float(row.get("credit_amount") or 0)

    for account, data in totals.items():
        cumulative_debit = data["opening_debit"] + data["period_debit"]
        cumulative_credit = data["opening_credit"] + data["period_credit"]
        data["cumulative_debit"] = cumulative_debit
        data["cumulative_credit"] = cumulative_credit
        net = cumulative_debit - cumulative_credit
        data["ending_debit"] = net if net > 0 else 0.0
        data["ending_credit"] = -net if net < 0 else 0.0

    return dict(totals)


def _resolve_source_path(file_url_or_name):
    file_url_or_name = _clean(file_url_or_name)
    if not file_url_or_name:
        raise ValueError("缺少上传文件路径。")
    if file_url_or_name.startswith(("/files/", "/private/files/")) or "/" not in file_url_or_name:
        path = get_file_path(file_url_or_name)
    else:
        path = file_url_or_name
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f"找不到上传文件：{file_url_or_name}")
    return path


def _company_name():
    try:
        company = frappe.defaults.get_user_default("Company") or frappe.db.get_single_value("Global Defaults", "default_company")
        return company or ""
    except Exception:
        return ""


def generate_financial_vouchers(file_url_or_name, ai_classifier=None):
    source_path = _resolve_source_path(file_url_or_name)
    transactions = parse_bank_transactions(source_path)
    voucher_rows, review_notes, ai_stats = build_voucher_rows(transactions, ai_classifier=ai_classifier)
    bank_context = build_bank_balance_context(transactions)
    trial_balance = build_trial_balance(voucher_rows, bank_context=bank_context)
    workbook_bytes, month_serial = build_financial_report_workbook(
        voucher_rows, trial_balance, transactions, company_name=_company_name(), bank_context=bank_context
    )

    source_stem = os.path.splitext(os.path.basename(source_path))[0]
    safe_stem = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff._-]+", "_", source_stem).strip("_") or "银行交易明细"
    output_name = f"{safe_stem}_财务凭证报表_{month_serial}.xlsx"
    file_doc = save_file(output_name, workbook_bytes, None, None, is_private=1)

    debit_total = sum(float(row.get("debit_amount") or 0) for row in voucher_rows)
    credit_total = sum(float(row.get("credit_amount") or 0) for row in voucher_rows)
    if round(debit_total - credit_total, 2) != 0:
        review_notes.append(f"借贷合计不平衡：借方 {debit_total:.2f}，贷方 {credit_total:.2f}")

    text = (
        f"已根据上传的银行交易明细生成财务凭证报表。\n\n"
        f"- 识别交易：{len(transactions)} 笔\n"
        f"- 生成凭证分录行：{len(voucher_rows)} 行\n"
        f"- 规则未命中候选：{ai_stats['candidate_count']} 条\n"
        f"- AI辅助采纳：{ai_stats['applied_count']} 条\n"
        f"- 借方合计：￥{debit_total:,.2f}\n"
        f"- 贷方合计：￥{credit_total:,.2f}\n"
        f"- 输出文件：{file_doc.file_name}\n"
    )
    if review_notes:
        preview = "\n".join(f"  - {note}" for note in review_notes[:10])
        more = f"\n  - 另有 {len(review_notes) - 10} 条复核提示，请打开 Excel 核对。" if len(review_notes) > 10 else ""
        text += f"\n以下交易建议人工复核：\n{preview}{more}\n"
    else:
        text += "\n所有凭证均已通过规则或 AI 辅助匹配，借贷平衡。\n"

    return {
        "text": text,
        "file_name": file_doc.file_name,
        "file_url": file_doc.file_url,
        "transaction_count": len(transactions),
        "voucher_row_count": len(voucher_rows),
        "review_count": len(review_notes),
        "ai_candidate_count": ai_stats["candidate_count"],
        "ai_applied_count": ai_stats["applied_count"],
        "ai_invalid_count": ai_stats["invalid_count"],
    }
